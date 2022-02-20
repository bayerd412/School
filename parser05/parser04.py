from bs4 import BeautifulSoup
import requests
import csv
from openpyxl import load_workbook
from rich.console import Console
import eel
       
#Все ссылки для поиска
items = [
    "http://gymnaz1-murm.ru/",
    "http://2gimn51.ru/",
    "http://gym3murmansk.ucoz.ru/publ/?page1",
    "http://gimn5.murm.eduru.ru/news",
    "http://gymn6.com.ru/",
    "http://gim7.murm.eduru.ru/",
    "http://www.gimnazia8.ru/",
    "https://mml.ucoz.org/",
    "http://www.mplmurmansk.ru/",
    "http://murmanprg24.ucoz.ru/",
    "http://prog40.ru/novosti",
    # "https://progimnaziya.murm.eduru.ru/news",
    # "http://progimnaziya61.murm.eduru.ru/news",
    # "http://mschool1.ru/",
    # "http://roslshk.moy.su/",
    # "https://murm-shkola4.murm.eduru.ru/news",
    # "http://mou-school11.ucoz.ru/",
    # "http://www.mou16-murmansk.ru/",
    # "http://my-school18.ucoz.ru/",
    # "http://kadet-murmansk.ru/news/",
    # "http://www.school20.com.ru/index.php/novosti-i-ob-yavleniya",
    # "https://school21.murm.eduru.ru/news",
    # "http://www.school22mur.ru/",
    # "http://school23mur.ucoz.ru/",
    # "http://school27-murman.my1.ru/index/novosti_i_objavlenija/0-136",
    # "http://skosh8.ucoz.ru/",
    # "http://school28.ucoz.ru/",
    # "https://skole31.murm.eduru.ru/news",
    # "http://murman-school33.ucoz.ru/",
    # "https://school34-murman.my1.ru/",
    # "http://school36.murmansk.su/",
    # "http://www.school42.znaet.ru/",
    # "http://sc43murm.com.ru/index.php/novosti",
    # "http://murman-school44.ru/",
    # "https://sosh45.murm.eduru.ru/news",
    # "https://school49.murm.eduru.ru/news",
    # "http://school50.su/",
    # "http://murmansk57.ru/",
    # "http://www.school58.org.ru/"
]

headers = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/95.0.4638.69 Safari/537.36'
}


searching_list = [] 
find_items = []
find_links = []
console = Console()


def save_file_txt(find_items, items, find_links):

    not_found_f = open(f'{searching_list[0]}_не_найдено.txt', "w")           
    found_f = open(f'{searching_list[0]}_найдено.txt', "w") 
    found_f_links = open(f'{searching_list[0]}_ссылки.txt', "w")

    for i in find_items:         
        found_f.write(i + '\n')
    for i in items:
        not_found_f.write(i + '\n')
    for i in find_links:
        found_f_links.write(i + '\n')
        
    found_f.close()
    not_found_f.close() 
    found_f_links.close()


def save_file_csv(find_items, items, find_links):
    
    with open(f'{searching_list[0]}_не_найдено.csv', 'a', encoding='utf_8_sig', newline='') as file:
        writer = csv.writer(file, delimiter=';')
        for item in items:
            writer.writerow([item])

    with open(f'{searching_list[0]}_найдено.csv', 'a', encoding='utf_8_sig', newline='') as file:
        writer = csv.writer(file, delimiter=';')
        for find_item in find_items:
            writer.writerow([find_item])

    with open(f'{searching_list[0]}_ссылки.csv', 'a', encoding='utf_8_sig', newline='') as file:
        writer = csv.writer(file, delimiter=';')
        for find_link in find_links:
            writer.writerow([find_link])
            

def choose_save_file_mode():
    global save_file_mode
    save_file_mode = input('\nВведите вид файла(csv/txt): ')

    if save_file_mode == 'csv':
        save_file_mode = 0
    elif save_file_mode == 'txt':
        save_file_mode = 1
    elif save_file_mode == 'exit':
        exit("Досвидания!")
    else:
        console.print('[red]Такого режима нет[/red]')
        choose_save_file_mode()


def ask_method_input_kewords():
    global save_searching_list_mode
    ask_save_searching_list_mode = input('Введите вид ввода слов(xlsx/клавиатура): ')

    if ask_save_searching_list_mode == 'xlsx':
        save_searching_list_mode = 0       
    elif ask_save_searching_list_mode == 'клавиатура':
        save_searching_list_mode = 1
    elif ask_save_searching_list_mode == 'exit':
        exit("Досвидания!")
    else:
        console.print('[red]Такого режима нет[/red]')
        ask_method_input_kewords()
    

def input_keywords_from_excel():
    try:
        global searching_list
        path = input('\nВведите название excel файла: ')
        sheet_number = input('\nВведите название листа: ')
        wb = load_workbook(f"{path}.xlsx")
        ws = wb[sheet_number]
        column = ws['A']
        searching_list = [column[x].value for x in range(len(column))]
        print(searching_list)
    except Exception:
        console.print('[red]Ошибка![/red]\n'
              '[red]Проверьте имя файла или листа[/red]')
        input_keywords_from_excel()

@eel.expose
def input_keywords(searching_text):                 
    searching_list.append(searching_text)
    return str(searching_list)
 
         
def delete_keywords_from_list():
    delete_item = input('\nВведите индекс слова для удаления(next для продолжения): ')
    try:
        if delete_item == 'next':
            return
        elif delete_item == 'exit':
            exit("Досвидания!")
        else:
            searching_list.pop(int(delete_item))
            print(f'Список слов для поиска: {searching_list}')
            delete_keywords_from_list()
    except Exception:
        console.print('[red]Ошибочный индекс[/red]')
        delete_keywords_from_list()
                   
# def parse_with_links(items, find_links, find_items):
#     print(items)
#     try:                         
#         for item in items:
#             req = requests.get(item)
#             src = req.text
            
#             soup = BeautifulSoup(src, 'lxml')
            
#             links = soup.find_all('a')
            
#             for link in links:
#                 req = requests.get(link, headers=headers)
#                 src = req.text
                
#                 soup = BeautifulSoup(src, 'lxml')
#                 hrefs = soup.find_all('a')
#                 site_text = soup.get_text().lower()
                
#                 break_flag = False
                
#                 for href in hrefs:
#                     href_text = href.get_text().lower()           
#                     for keyword in searching_list:
#                         if keyword in href_text:
#                             break_flag = True
#                             if item in href.get('href'):
#                                 find_links.append(href.get('href'))
#                             else:
#                                 find_links.append(item + href.get('href'))
#                             find_items.append(item)
#                             items.remove(item)
#                             break
#                         elif break_flag == False and href == hrefs[-1] and keyword in site_text:
#                             break_flag = True
#                             find_links.append(item)
#                             find_items.append(item)
#                             items.remove(item)
#                             break
                        
#                     if break_flag:
#                         break
#             if break_flag:
#                 break 
                  
#         if save_file_mode == 0:
#             save_file_csv(find_items, items, find_links)
#         else:
#             save_file_txt(find_items, items, find_links)

#     except Exception:
#         pass
            
                        
def parse():
    #Использовал копию списка, иначе возникает баг и читается только половина ссылок.
    for item in items.copy():
        req = requests.get(item, headers=headers)
        src = req.text
        
        soup = BeautifulSoup(src, 'lxml')
        hrefs = soup.find_all('a')
        site_text = soup.get_text().lower()
        
        break_flag = False
        for href in hrefs:
            href_text = href.get_text().lower()           
            for keyword in searching_list:
                if keyword in href_text:
                    break_flag = True
                    if item in href.get('href'):
                        find_links.append(href.get('href'))
                    else:
                        find_links.append(item + href.get('href'))
                    find_items.append(item)
                    items.remove(item)
                    break
                elif break_flag == False and href == hrefs[-1] and keyword in site_text:
                    break_flag = True
                    find_links.append(item)
                    find_items.append(item)
                    items.remove(item)
                    break
                
            if break_flag:
                break
                    
                                          
# def main():
    
#     eel.init('web')
#     eel.start('main.html', size=(400, 600))
    
    # console.print("[green]Для выхода из парсера введите: exit[/green]")
    # console.print("[green]Для перехода к следующему шагу введите: next \n[/green]")
    # ask_method_input_kewords()
    # if save_searching_list_mode == 0:
    #     input_keywords_from_excel()
    # else:
    # input_keywords()
        
    # delete_keywords_from_list()
    # choose_save_file_mode()
    # parse() 
    
    # if save_file_mode == 0:
    #     save_file_csv(find_items, items, find_links)
    # else:
    #     save_file_txt(find_items, items, find_links)
          
#     main()
       
# main()

eel.init('web')
eel.start('main.html', size=(400, 600))